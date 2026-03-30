#!/usr/bin/env python3
"""
estudio_ruido_nom.py
════════════════════════════════════════════════════════════════════════
Automatización de Estudios de Ruido de Fuente Fija
Norma Oficial Mexicana NOM-081-SEMARNAT-1994

Uso como módulo:
    from estudio_ruido_nom import cargar_csv, EstudioRuidoNOM
    datos_f = cargar_csv("fuente.csv")
    datos_b = cargar_csv("fondo.csv")
    estudio = EstudioRuidoNOM(datos_f, datos_b, {"Sitio": "ZC1"})
    estudio.calcular()
    print(estudio.reporte())
    estudio.exportar_excel("resultado.xlsx")

Uso como CLI:
    python estudio_ruido_nom.py --template
    python estudio_ruido_nom.py --fuente fuente.csv --fondo fondo.csv
    python estudio_ruido_nom.py --fuente fuente.csv --fondo fondo.csv \\
        --meta "Sitio=ZC1 Honda" "Fecha=2024-03-15" --output resultado.xlsx
════════════════════════════════════════════════════════════════════════
"""

import argparse
import math
import os
import sys
from pathlib import Path

# ── Verificar dependencias ──────────────────────────────────────────
try:
    import numpy as np
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(
        f"\n[ERROR] Librería faltante: {e}\n"
        f"Instala las dependencias con:\n"
        f"  pip install pandas numpy openpyxl\n"
    )

# ╔══════════════════════════════════════════════════════════════════╗
# ║  CONSTANTES NOM-081-SEMARNAT-1994                               ║
# ╚══════════════════════════════════════════════════════════════════╝

PERIODOS_FUENTE   = ["A", "B", "C", "D", "E"]
PERIODOS_FONDO    = ["I", "II", "III", "IV", "V"]
N_LECTURAS        = 35
K_N10             = 1.2817   # percentil 90 de distribución normal estándar
K_CE              = 0.9023   # factor de corrección por valores extremos
UMBRAL_DELTA      = 0.75     # umbral Δ50 para aplicar corrección por fondo


# ╔══════════════════════════════════════════════════════════════════╗
# ║  FUNCIONES UTILITARIAS                                          ║
# ╚══════════════════════════════════════════════════════════════════╝

def cargar_csv(path: str) -> dict:
    """
    Carga y valida un archivo CSV de lecturas de ruido.

    Formato esperado
    ----------------
    Columna 'lectura': índice 1-35
    Columnas de periodos: A,B,C,D,E (fuente) o I,II,III,IV,V (fondo)
    Todos los valores deben ser numéricos (dB).

    Returns
    -------
    dict con claves:
        'df'       → pandas DataFrame limpio
        'periodos' → lista de columnas de periodo
        'archivo'  → ruta del archivo cargado
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {path}")
    if path.suffix.lower() not in (".csv", ".txt"):
        raise ValueError(f"Se esperaba un archivo .csv, se recibió: {path.suffix}")

    df = pd.read_csv(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]

    if "lectura" not in df.columns:
        raise ValueError(
            f"'{path.name}' debe tener una columna llamada 'lectura'.\n"
            f"  Columnas encontradas: {list(df.columns)}"
        )

    periodos = [c for c in df.columns if c != "lectura"]
    if not periodos:
        raise ValueError(f"'{path.name}' no contiene columnas de periodos.")

    # Convertir a numérico
    for p in periodos:
        df[p] = pd.to_numeric(df[p], errors="coerce")

    if df[periodos].isna().any().any():
        celdas_malas = df[periodos].isna().stack()
        celdas_malas = celdas_malas[celdas_malas].index.tolist()
        raise ValueError(
            f"Valores no numéricos en '{path.name}': {celdas_malas[:5]}"
        )

    if len(df) != N_LECTURAS:
        raise ValueError(
            f"'{path.name}' tiene {len(df)} filas; se esperan {N_LECTURAS}."
        )

    return {"df": df, "periodos": periodos, "archivo": str(path)}


def generar_templates(directorio: str = ".") -> None:
    """
    Crea archivos CSV plantilla listos para llenar con datos de campo.

    Archivos generados
    ------------------
    fuente_template.csv  →  columnas: lectura, A, B, C, D, E
    fondo_template.csv   →  columnas: lectura, I, II, III, IV, V
    """
    dir_path = Path(directorio)
    dir_path.mkdir(parents=True, exist_ok=True)

    for nombre, periodos in [
        ("fuente_template.csv", PERIODOS_FUENTE),
        ("fondo_template.csv",  PERIODOS_FONDO),
    ]:
        df = pd.DataFrame({"lectura": range(1, N_LECTURAS + 1)})
        for p in periodos:
            df[p] = ""
        out = dir_path / nombre
        df.to_csv(out, index=False)
        print(f"[OK] Plantilla creada: {out.resolve()}")


# ╔══════════════════════════════════════════════════════════════════╗
# ║  FUNCIONES DE CÁLCULO (internas)                                ║
# ╚══════════════════════════════════════════════════════════════════╝

def _calcular_periodo(lecturas: np.ndarray) -> dict:
    """
    Calcula indicadores acústicos para un vector de lecturas.

    Fórmulas NOM-081
    ----------------
    N50  = media aritmética
    σ    = desviación estándar muestral (ddof=1)
    N10  = N50 + 1.2817 × σ
    Neq  = 10 × log₁₀[ (1/n) × Σ 10^(Nᵢ/10) ]
    """
    n = len(lecturas)
    n50   = float(np.mean(lecturas))
    sigma = float(np.std(lecturas, ddof=1))
    n10   = n50 + K_N10 * sigma
    neq   = 10.0 * math.log10(
        (1.0 / n) * float(np.sum(10.0 ** (lecturas / 10.0)))
    )
    return {"N50": n50, "sigma": sigma, "N10": n10, "Neq": neq}


def _calcular_bloque(df: pd.DataFrame, periodos: list) -> dict:
    """
    Calcula estadísticos por periodo y promedios globales para un bloque
    (fuente o fondo).

    Promedios globales
    ------------------
    N50_prom  = media aritmética de los N50 por periodo
    σ_prom    = media aritmética de las σ por periodo
    N10_prom  = N50_prom + 1.2817 × σ_prom
    (Neq)eq   = 10 × log₁₀[ (1/5) × Σ 10^(Neqᵢ/10) ]  ← promedio energético
    """
    stats_periodos = {}
    for p in periodos:
        stats_periodos[p] = _calcular_periodo(df[p].values.astype(float))

    n50s   = [stats_periodos[p]["N50"]   for p in periodos]
    sigmas = [stats_periodos[p]["sigma"] for p in periodos]
    neqs   = [stats_periodos[p]["Neq"]   for p in periodos]

    n50_prom   = float(np.mean(n50s))
    sigma_prom = float(np.mean(sigmas))
    n10_prom   = n50_prom + K_N10 * sigma_prom
    neq_eq     = 10.0 * math.log10(
        (1.0 / len(periodos)) * sum(10.0 ** (n / 10.0) for n in neqs)
    )

    return {
        "periodos":    stats_periodos,
        "N50_prom":    n50_prom,
        "sigma_prom":  sigma_prom,
        "N10_prom":    n10_prom,
        "Neq_eq":      neq_eq,
    }


# ╔══════════════════════════════════════════════════════════════════╗
# ║  CLASE PRINCIPAL                                                ║
# ╚══════════════════════════════════════════════════════════════════╝

class EstudioRuidoNOM:
    """
    Realiza el cálculo completo de un estudio de ruido de fuente fija
    conforme a la NOM-081-SEMARNAT-1994.

    Parámetros
    ----------
    datos_fuente : dict  →  resultado de cargar_csv() para archivo de fuente
    datos_fondo  : dict  →  resultado de cargar_csv() para archivo de fondo
    metadata     : dict  →  información descriptiva del estudio (opcional)
                            Ej: {"Sitio": "ZC1 Honda", "Fecha": "2024-03-15",
                                 "Responsable": "Ing. García"}
    """

    def __init__(
        self,
        datos_fuente: dict,
        datos_fondo:  dict,
        metadata:     dict = None,
    ):
        self.df_fuente  = datos_fuente["df"]
        self.df_fondo   = datos_fondo["df"]
        self.per_fuente = datos_fuente["periodos"]
        self.per_fondo  = datos_fondo["periodos"]
        self.metadata   = metadata or {}
        self._res       = None   # cache de resultados

    # ── API pública ─────────────────────────────────────────────────

    def calcular(self) -> dict:
        """
        Ejecuta todos los cálculos según la secuencia NOM-081.

        Returns
        -------
        dict con claves:
            fuente, fondo       → resultados por bloque (ver _calcular_bloque)
            Ce                  → corrección por valores extremos
            delta50             → diferencia N50 fuente − fondo
            N50_corr            → N'50 = N50_prom_fuente + Ce
            Cf_aplica           → bool, si Δ50 ≥ 0.75
            Cf                  → corrección por fondo (None si no aplica)
            Nff                 → max(N'50, (Neq)eq)
            Nff_corr            → (N')ff con Cf aplicada (o igual a Nff)
            metadata            → información del estudio
        """
        fuente = _calcular_bloque(self.df_fuente, self.per_fuente)
        fondo  = _calcular_bloque(self.df_fondo,  self.per_fondo)

        # ── Corrección por valores extremos ──
        ce = K_CE * fuente["sigma_prom"]

        # ── N'50 ──
        delta50  = fuente["N50_prom"] - fondo["N50_prom"]
        n50_corr = fuente["N50_prom"] + ce

        # ── Corrección por fondo ──
        if delta50 >= UMBRAL_DELTA:
            cf_aplica = True
            cf = -(delta50 + 9.0) + 3.0 * math.sqrt(4.0 * delta50 - 3.0)
        else:
            cf_aplica = False
            cf = None

        # ── Nivel de Fuente Fija ──
        # Ce/N'50 solo aplica cuando fuente >= fondo (criterio de práctica profesional).
        # Cuando Δ50 < 0 (fondo > fuente) la variabilidad no es atribuible a la fuente.
        n50_para_max = n50_corr if delta50 >= 0 else fuente["N50_prom"]
        nff      = max(n50_para_max, fuente["Neq_eq"])
        nff_corr = nff + cf if cf_aplica else nff

        self._res = {
            "fuente":    fuente,
            "fondo":     fondo,
            "Ce":        ce,
            "delta50":   delta50,
            "N50_corr":  n50_corr,
            "Cf_aplica": cf_aplica,
            "Cf":        cf,
            "Nff":       nff,
            "Nff_corr":  nff_corr,
            "metadata":  self.metadata,
        }
        return self._res

    def reporte(self) -> str:
        """Genera un reporte completo formateado para consola."""
        if self._res is None:
            self.calcular()
        r = self._res

        SEP  = "═" * 68
        SEP2 = "─" * 68
        fmt  = lambda v, d=2: f"{v:.{d}f}" if v is not None else "N/A"

        lines = [SEP]
        lines.append("  ESTUDIO DE RUIDO DE FUENTE FIJA — NOM-081-SEMARNAT-1994")
        lines.append(SEP)

        if self.metadata:
            for k, v in self.metadata.items():
                lines.append(f"  {k:<20}: {v}")
            lines.append(SEP2)

        def tabla_bloque(titulo, bloque, periodos):
            lines.append(f"\n{SEP2}")
            lines.append(f"  {titulo}")
            lines.append(SEP2)
            w = 9  # ancho por columna
            # Encabezado
            hdr = f"  {'Indicador':<14}" + "".join(f"{p:>{w}}" for p in periodos)
            hdr += f"{'PROMEDIO':>{w+2}}"
            lines.append(hdr)
            lines.append(f"  {'─'*62}")

            defs = [
                ("N50  (dB)", "N50"),
                ("σ    (dB)", "sigma"),
                ("N10  (dB)", "N10"),
                ("Neq  (dB)", "Neq"),
            ]
            proms = {
                "N50":   bloque["N50_prom"],
                "sigma": bloque["sigma_prom"],
                "N10":   bloque["N10_prom"],
                "Neq":   bloque["Neq_eq"],
            }
            for label, key in defs:
                vals = [bloque["periodos"][p][key] for p in periodos]
                row = f"  {label:<14}" + "".join(f"{v:>{w}.2f}" for v in vals)
                row += f"{proms[key]:>{w+2}.2f}"
                lines.append(row)

        tabla_bloque("RUIDO DE FUENTE  (Periodos A–E)", r["fuente"], self.per_fuente)
        tabla_bloque("RUIDO DE FONDO   (Periodos I–V)", r["fondo"],  self.per_fondo)

        lines.append(f"\n{SEP2}")
        lines.append("  CORRECCIONES Y RESULTADO FINAL")
        lines.append(SEP2)

        def fila_r(label, valor, unidad="dB"):
            if isinstance(valor, str):
                lines.append(f"  {label:<38}  {valor}")
            else:
                lines.append(f"  {label:<38}  {fmt(valor)} {unidad}")

        fila_r("Ce  — Correc. por valores extremos",  r["Ce"])
        fila_r("Δ50 — Diferencia fuente − fondo",     r["delta50"])
        fila_r("N'50 = N50_prom + Ce",                r["N50_corr"])

        if r["Cf_aplica"]:
            fila_r("Cf  — Correc. por fondo",         r["Cf"])
        else:
            fila_r("Cf  — Correc. por fondo",
                   f"No Aplica  (Δ50 = {fmt(r['delta50'])} < {UMBRAL_DELTA})")

        lines.append("")
        lines.append(
            f"  Nff  = max(N'50, (Neq)eq) = "
            f"max({fmt(r['N50_corr'])}, {fmt(r['fuente']['Neq_eq'])}) "
            f"= {fmt(r['Nff'])} dB"
        )
        lines.append(f"  (N')ff = {fmt(r['Nff_corr'])} dB")
        lines.append(SEP)

        return "\n".join(lines)

    def exportar_excel(self, path: str) -> None:
        """
        Exporta los resultados a un archivo Excel con tres hojas:
        'Fuente', 'Fondo' y 'Resumen'.
        """
        if self._res is None:
            self.calcular()

        r  = self._res
        wb = openpyxl.Workbook()

        # ── Paleta de colores ─────────────────────────────────────
        C = {
            "azul_oscuro":  "1F4E79",
            "azul_medio":   "2E75B6",
            "azul_claro":   "D6E4F0",
            "azul_header":  "4472C4",
            "verde_claro":  "E2EFDA",
            "amarillo":     "FFF2CC",
            "gris_fila":    "F2F2F2",
            "blanco":       "FFFFFF",
        }

        thin   = Side(style="thin")
        thick  = Side(style="medium")
        border_thin   = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_bottom = Border(bottom=thick)

        # ── Helpers de estilo ─────────────────────────────────────

        def _fill(color: str) -> PatternFill:
            return PatternFill("solid", fgColor=color)

        def _font(bold=False, color="000000", size=10) -> Font:
            return Font(bold=bold, color=color, size=size)

        def _align(h="center", v="center", wrap=False) -> Alignment:
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def write(ws, row, col, value, bold=False, fgcolor=None,
                  num_fmt=None, halign="center", size=10, fgfont=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(bold=bold, size=size,
                               color=fgfont or "000000")
            c.alignment = _align(halign)
            if fgcolor:
                c.fill = _fill(fgcolor)
            if num_fmt:
                c.number_format = num_fmt
            return c

        def merge_title(ws, row, c1, c2, text,
                        bg="1F4E79", fg="FFFFFF", size=11, h=22):
            ws.merge_cells(
                start_row=row, start_column=c1,
                end_row=row,   end_column=c2
            )
            c = ws.cell(row=row, column=c1, value=text)
            c.fill      = _fill(bg)
            c.font      = Font(bold=True, color=fg, size=size)
            c.alignment = _align("center")
            ws.row_dimensions[row].height = h
            return c

        def border_range(ws, r1, r2, c1, c2):
            for row in ws.iter_rows(min_row=r1, max_row=r2,
                                     min_col=c1, max_col=c2):
                for cell in row:
                    cell.border = border_thin

        # ══════════════════════════════════════════════════════════
        # Hoja de datos (Fuente o Fondo)
        # ══════════════════════════════════════════════════════════

        def crear_hoja_datos(nombre, df_data, periodos, bloque):
            ws = wb.active if nombre == "Fuente" else wb.create_sheet(nombre)
            ws.title = nombre
            ncol     = len(periodos)

            # ── Título principal ──
            merge_title(ws, 1, 1, ncol + 2,
                        f"RUIDO DE {nombre.upper()} — NOM-081-SEMARNAT-1994",
                        bg=C["azul_oscuro"], size=12, h=26)

            # ── Encabezados de columnas ──
            write(ws, 2, 1, "N° Lectura",
                  bold=True, fgcolor=C["azul_medio"], fgfont="FFFFFF")
            for i, p in enumerate(periodos, start=2):
                write(ws, 2, i, f"Periodo {p}",
                      bold=True, fgcolor=C["azul_medio"], fgfont="FFFFFF")

            ws.row_dimensions[2].height = 18

            # ── Lecturas ──
            for idx, row_data in df_data.iterrows():
                fila = idx + 3
                write(ws, fila, 1, int(row_data["lectura"]))
                bg = C["gris_fila"] if idx % 2 == 0 else None
                for j, p in enumerate(periodos, start=2):
                    write(ws, fila, j, float(row_data[p]),
                          num_fmt="0.0", fgcolor=bg)

            border_range(ws, 2, N_LECTURAS + 2, 1, ncol + 1)

            # ── Sección de resultados por periodo ──
            fila_ini = N_LECTURAS + 4

            merge_title(ws, fila_ini - 1, 1, ncol + 2,
                        "Resultados por Periodo",
                        bg=C["azul_medio"], size=10, h=18)

            # sub-encabezado
            write(ws, fila_ini, 1, "Indicador",
                  bold=True, fgcolor=C["azul_claro"])
            for i, p in enumerate(periodos, start=2):
                write(ws, fila_ini, i, p,
                      bold=True, fgcolor=C["azul_claro"])
            write(ws, fila_ini, ncol + 2, "PROMEDIO",
                  bold=True, fgcolor=C["azul_claro"])
            ws.row_dimensions[fila_ini].height = 16

            labels = [
                ("N50  (dB)", "N50",   bloque["N50_prom"]),
                ("σ    (dB)", "sigma", bloque["sigma_prom"]),
                ("N10  (dB)", "N10",   bloque["N10_prom"]),
                ("Neq  (dB)", "Neq",   bloque["Neq_eq"]),
            ]

            for off, (label, key, prom) in enumerate(labels):
                f = fila_ini + 1 + off
                bg_row = "EBF3FA" if off % 2 == 0 else C["blanco"]
                write(ws, f, 1, label, bold=True,
                      fgcolor=bg_row, halign="left")
                for j, p in enumerate(periodos, start=2):
                    write(ws, f, j,
                          round(bloque["periodos"][p][key], 2),
                          num_fmt="0.00", fgcolor=bg_row)
                write(ws, f, ncol + 2, round(prom, 2),
                      bold=True, num_fmt="0.00",
                      fgcolor=C["azul_claro"])

            border_range(ws, fila_ini, fila_ini + len(labels), 1, ncol + 2)

            # ── Anchos de columna ──
            ws.column_dimensions["A"].width = 14
            for i in range(2, ncol + 3):
                ws.column_dimensions[get_column_letter(i)].width = 13

        crear_hoja_datos("Fuente", self.df_fuente, self.per_fuente, r["fuente"])
        crear_hoja_datos("Fondo",  self.df_fondo,  self.per_fondo,  r["fondo"])

        # ══════════════════════════════════════════════════════════
        # Hoja Resumen
        # ══════════════════════════════════════════════════════════

        ws = wb.create_sheet("Resumen")

        merge_title(ws, 1, 1, 4,
                    "RESUMEN EJECUTIVO — NOM-081-SEMARNAT-1994",
                    bg=C["azul_oscuro"], size=13, h=30)

        fila = 3

        # ── Metadata ──────────────────────────────────────────────
        if self.metadata:
            merge_title(ws, fila, 1, 4,
                        "INFORMACIÓN DEL ESTUDIO",
                        bg=C["azul_header"], size=10, h=18)
            fila += 1
            for k, v in self.metadata.items():
                write(ws, fila, 1, k, bold=True,
                      fgcolor="EBF3FA", halign="left")
                ws.merge_cells(
                    start_row=fila, start_column=2,
                    end_row=fila,   end_column=4
                )
                ws.cell(row=fila, column=2,
                        value=str(v)).alignment = _align("left")
                fila += 1
            border_range(ws, fila - len(self.metadata) - 1,
                         fila - 1, 1, 4)
            fila += 1

        # ── Tabla comparativa ────────────────────────────────────
        merge_title(ws, fila, 1, 4,
                    "COMPARATIVO FUENTE vs FONDO",
                    bg=C["azul_medio"], size=10, h=18)
        fila += 1

        for j, h in enumerate(["Indicador", "Fuente (A–E)",
                                "Fondo (I–V)", "Δ (F−B)"], start=1):
            write(ws, fila, j, h, bold=True,
                  fgcolor=C["azul_header"], fgfont="FFFFFF")
        fila += 1

        comps = [
            ("N50 promedio (dB)",
             r["fuente"]["N50_prom"],  r["fondo"]["N50_prom"]),
            ("σ promedio (dB)",
             r["fuente"]["sigma_prom"], r["fondo"]["sigma_prom"]),
            ("N10 promedio (dB)",
             r["fuente"]["N10_prom"],  r["fondo"]["N10_prom"]),
            ("(Neq)eq (dB)",
             r["fuente"]["Neq_eq"],    r["fondo"]["Neq_eq"]),
        ]
        ini_comp = fila
        for off, (lbl, vf, vb) in enumerate(comps):
            bg = "EBF3FA" if off % 2 == 0 else C["blanco"]
            write(ws, fila, 1, lbl, bold=True,
                  fgcolor=bg, halign="left")
            write(ws, fila, 2, round(vf, 2),
                  num_fmt="0.00", fgcolor=bg)
            write(ws, fila, 3, round(vb, 2),
                  num_fmt="0.00", fgcolor=bg)
            diff = round(vf - vb, 2)
            fg_diff = C["verde_claro"] if diff >= 0 else C["amarillo"]
            write(ws, fila, 4, diff,
                  bold=True, num_fmt="+0.00;-0.00;0.00",
                  fgcolor=fg_diff)
            fila += 1
        border_range(ws, ini_comp - 1, fila - 1, 1, 4)
        fila += 1

        # ── Correcciones ─────────────────────────────────────────
        merge_title(ws, fila, 1, 4,
                    "CORRECCIONES",
                    bg=C["azul_medio"], size=10, h=18)
        fila += 1

        corrs = [
            ("Ce — Correc. por valores extremos (0.9023 × σ_prom)",
             f"{r['Ce']:.2f} dB"),
            ("Δ50 — Diferencia N50 fuente − fondo",
             f"{r['delta50']:.2f} dB"),
            ("N'50 — N50 corregido (N50_prom + Ce)",
             f"{r['N50_corr']:.2f} dB"),
            ("Cf — Corrección por fondo",
             f"{r['Cf']:.2f} dB" if r["Cf_aplica"]
             else f"No Aplica  (Δ50={r['delta50']:.2f} < {UMBRAL_DELTA})"),
        ]
        ini_corr = fila
        for off, (lbl, val) in enumerate(corrs):
            bg = "EBF3FA" if off % 2 == 0 else C["blanco"]
            ws.merge_cells(
                start_row=fila, start_column=1,
                end_row=fila,   end_column=3
            )
            write(ws, fila, 1, lbl, bold=False,
                  fgcolor=bg, halign="left")
            fg_val = (C["amarillo"]
                      if not r["Cf_aplica"] and lbl.startswith("Cf")
                      else C["azul_claro"])
            write(ws, fila, 4, val, bold=True,
                  fgcolor=fg_val)
            fila += 1
        border_range(ws, ini_corr - 1, fila - 1, 1, 4)
        fila += 1

        # ── Resultado final ───────────────────────────────────────
        merge_title(ws, fila, 1, 4,
                    "NIVEL DE FUENTE FIJA — RESULTADO",
                    bg=C["azul_oscuro"], size=11, h=22)
        fila += 1

        res_finales = [
            (f"Nff = max(N'50={r['N50_corr']:.2f}, (Neq)eq={r['fuente']['Neq_eq']:.2f})",
             f"{r['Nff']:.2f} dB"),
            ("(N')ff — Nivel final con todas las correcciones",
             f"{r['Nff_corr']:.2f} dB"),
        ]
        ini_fin = fila
        for lbl, val in res_finales:
            ws.merge_cells(
                start_row=fila, start_column=1,
                end_row=fila,   end_column=3
            )
            c1 = ws.cell(row=fila, column=1, value=lbl)
            c1.font      = Font(bold=True, size=11)
            c1.alignment = _align("left")
            c1.fill      = _fill(C["azul_claro"])
            c4 = ws.cell(row=fila, column=4, value=val)
            c4.font      = Font(bold=True, size=13, color=C["azul_oscuro"])
            c4.alignment = _align("center")
            c4.fill      = _fill(C["verde_claro"])
            ws.row_dimensions[fila].height = 24
            fila += 1
        border_range(ws, ini_fin - 1, fila - 1, 1, 4)

        # ── Anchos hoja Resumen ───────────────────────────────────
        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 20

        # ── Guardar ───────────────────────────────────────────────
        wb.save(path)
        print(f"[OK] Excel exportado: {Path(path).resolve()}")


# ╔══════════════════════════════════════════════════════════════════╗
# ║  CLI — Interfaz de Línea de Comandos                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="estudio_ruido_nom",
        description=(
            "Cálculo de ruido de fuente fija según NOM-081-SEMARNAT-1994\n"
            "──────────────────────────────────────────────────────────────\n"
            "Genera plantillas:  python estudio_ruido_nom.py --template\n"
            "Calcular:           python estudio_ruido_nom.py --fuente fuente.csv --fondo fondo.csv\n"
            "Exportar Excel:     python estudio_ruido_nom.py --fuente fuente.csv --fondo fondo.csv\n"
            "                       --output resultado.xlsx"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--fuente", metavar="CSV",
        help="Ruta al CSV de ruido de fuente (columnas: lectura, A, B, C, D, E)"
    )
    parser.add_argument(
        "--fondo", metavar="CSV",
        help="Ruta al CSV de ruido de fondo (columnas: lectura, I, II, III, IV, V)"
    )
    parser.add_argument(
        "--output", metavar="XLSX",
        help="Ruta de salida del archivo Excel (opcional)"
    )
    parser.add_argument(
        "--template", action="store_true",
        help="Genera plantillas CSV vacías y sale"
    )
    parser.add_argument(
        "--template-dir", metavar="DIR", default=".",
        help="Directorio para las plantillas (default: directorio actual)"
    )
    parser.add_argument(
        "--meta", nargs="+", metavar="CLAVE=VALOR",
        help=(
            "Metadatos del estudio (pares CLAVE=VALOR).\n"
            "Ejemplo: --meta Sitio='ZC1 Honda' Fecha=2024-03-15 Responsable='Ing. García'"
        )
    )
    return parser


def main():
    parser = _build_parser()
    args   = parser.parse_args()

    # ── Modo: generar plantillas ──────────────────────────────────
    if args.template:
        generar_templates(args.template_dir)
        return

    # ── Validar argumentos requeridos ─────────────────────────────
    if not args.fuente or not args.fondo:
        parser.print_help()
        print(
            "\n[ERROR] Se requieren --fuente y --fondo.\n"
            "  Usa --template para generar plantillas CSV vacías.\n"
        )
        sys.exit(1)

    # ── Cargar datos ──────────────────────────────────────────────
    try:
        datos_fuente = cargar_csv(args.fuente)
        datos_fondo  = cargar_csv(args.fondo)
    except (FileNotFoundError, ValueError) as e:
        sys.exit(f"\n[ERROR] {e}\n")

    # ── Parsear metadata ──────────────────────────────────────────
    metadata = {}
    if args.meta:
        for item in args.meta:
            if "=" in item:
                k, v = item.split("=", 1)
                metadata[k.strip()] = v.strip()
            else:
                print(f"[AVISO] Ignorando metadato malformado: '{item}' (formato: CLAVE=VALOR)")

    # ── Calcular ──────────────────────────────────────────────────
    estudio = EstudioRuidoNOM(datos_fuente, datos_fondo, metadata)
    estudio.calcular()

    # ── Mostrar reporte en consola ────────────────────────────────
    print(estudio.reporte())

    # ── Exportar Excel ────────────────────────────────────────────
    if args.output:
        try:
            estudio.exportar_excel(args.output)
        except PermissionError:
            sys.exit(
                f"\n[ERROR] No se puede escribir en '{args.output}'.\n"
                f"  Verifica que el archivo no esté abierto en Excel.\n"
            )
        except Exception as e:
            sys.exit(f"\n[ERROR] al exportar Excel: {e}\n")


if __name__ == "__main__":
    main()